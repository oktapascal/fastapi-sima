from typing import Annotated, Optional
from datetime import datetime
import numpy as np
import math
import time
import os
from typing import List
import pandas
import pyodbc
from dotenv import load_dotenv
from fastapi.middleware.cors import CORSMiddleware
from fastapi import FastAPI, BackgroundTasks, File, UploadFile, Form, Body, Query
from fastapi.responses import FileResponse
from openpyxl.styles import NamedStyle, PatternFill, Alignment
from scipy import stats
from pydantic import BaseModel

from database import database

load_dotenv()


def connect_dbsima():
    engine = database.Database(os.getenv('DB_USER1'), os.getenv('DB_PASSWORD1'), os.getenv('DB_HOST1'),
                               os.getenv('DB_NAME1'))

    dbsima_engine = engine.connect()

    return dbsima_engine


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST"],
    allow_headers=["*"]
)

"""
A function that serves as the root endpoint for the FastAPI application.
This function returns a dictionary with the keys 'status' and 'message',
representing the status and message of the API response.

Parameters:
None

Returns:
A dictionary with the keys 'status' and 'message'.
"""


@app.get('/')
def root():
    return {'status': 'OK', 'message': 'Hello From Fastapi-SIMA'}


@app.get('/test-db')
def test_db():
    try:
        dbsima = connect_dbsima()
        print("{c} is working".format(c=dbsima))
        dbsima.close()

        return {'status': 'OK', 'message': 'Success Connect Database'}
    except pyodbc.Error as ex:
        print("{c} is not working".format(c=dbsima))


@app.get('/api/export/excel/perangkat')
def export_excel_perangkat(background_task: BackgroundTasks, kode_jenis: str | None = None,
                           kode_lokasi: str | None = None, tahun: str | None = None, kode_area: str | None = None,
                           kode_fm: str | None = None, kode_bm: str | None = None, kode_ktg: str | None = None,
                           kode_subktg: str | None = None, status_aktif: str | None = None):
    columns = ["no", "id_regional", "id_area", "nama_area", "id_bm", "nama_bm", "id_witel", "nama_witel", "id_location",
               "nama_lokasi", "id_gedung", "nama_gedung", "id_room", "id_lantai",
               "nama_lantai", "id_jenis", "nama_jenis", "id_kategori", "nama_kategori", "id_subkategori",
               "nama_subkategori", "nama_perangkat", "is_ceklis", "kode_merk", "nama_merk", "satuan", "jumlah",
               "kapasitas", "no_seri", "tipe", "tahun", "kondisi", "kode_milik", "nama_milik", "keterangan",
               "id_perangkat", "status_aktif", "tanggal_periksa", "nik_input", "updated_at"]

    where = "where e.flag_aktif <> '3'"
    if kode_jenis is not None and kode_jenis != "" and kode_jenis != "null":
        where = where + f"and a.kode_jenis in ({kode_jenis})"

    if kode_lokasi is not None and kode_lokasi != "" and kode_lokasi != "null":
        where = where + f"and a.kode_lokasi in ({kode_lokasi})"

    if tahun is not None and tahun != "" and tahun != "null":
        where = where + f"and a.tahun in ({tahun})"

    if kode_area is not None and kode_area != "" and kode_area != "null":
        where = where + f"and e.kode_area in ({kode_area})"

    if kode_fm is not None and kode_fm != "" and kode_fm != "null":
        where = where + f"and a.kode_fm in ({kode_fm})"

    if kode_bm is not None and kode_bm != "" and kode_bm != "null":
        where = where + f"and a.kode_bm in ({kode_bm})"

    if kode_ktg is not None and kode_ktg != "" and kode_ktg != "null":
        where = where + f"and a.kode_kategori in ({kode_ktg})"

    if kode_subktg is not None and kode_subktg != "" and kode_subktg != "null":
        where = where + f"and a.kode_subkategori in ({kode_subktg})"

    if status_aktif is not None and status_aktif != "" and status_aktif != "null":
        where = where + f"and isnull(a.status_aktif, '1') in ({status_aktif})"

    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        cursor.execute(f"""
        select a.id no_perangkat, e.kode_area, a.kode_fm, l.nama nama_fm, a.kode_bm, m.nama nama_bm, a.kode_witel, k.nama nama_witel, a.kode_lokasi, d.nama_lokasi,
        a.kode_gedung, e.nama_gedung, a.kode_room, a.kode_lantai, g.nama_lantai, a.kode_jenis, h.nama_jenis,
        a.kode_kategori, i.nama_kategori, a.kode_subkategori, j.nama_sub_kategori, a.nama_perangkat, a.is_ceklis, a.kode_merk, n.nama_merk, a.satuan, a.jumlah,
        a.kapasitas, a.no_seri, a.model, a.tahun, a.kondisi, a.kode_milik, o.nama nama_milik, a.keterangan, concat(a.prefix,'-',a.indeks) id_perangkat, 
        case when isnull(a.status_aktif, '1') = '1' then 'ACTIVE' else 'INACTIVE' end status_aktif, a.kondisi_terakhir, a.nik_user, a.tgl_input
        from dev_am_perangkat a
        inner join am_gedung as e ON e.kode_gedung = a.kode_gedung and e.kode_lokasi='11'
        inner join am_locations as d ON a.kode_lokasi = d.id
        inner join gsd_rooms as f ON a.kode_room = f.id
        inner join am_floors as g ON a.kode_lantai = g.id
        left join am_perangkat_jenis as h ON a.kode_jenis = h.jenis_id
        left join am_perangkat_kategori as i ON a.kode_kategori = i.kategori_id
        left join am_perangkat_sub_kategori as j ON a.kode_subkategori = j.sub_kategori_id
        left join am_witel as k ON e.kode_witel = k.kode_witel
        left join am_fm l on a.kode_fm=l.kode_fm
        left join am_bm m on a.kode_bm=m.kode_bm
        left join am_perangkat_merk n on a.kode_merk=n.kode_merk
        left join am_milik o on a.kode_milik=o.kode_milik
        {where}
      """)

        dataframe = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns)

        today = datetime.today()
        unique_id = today.strftime('%Y%m%d%H%M%S')

        file_name = f'DATA_PERANGKAT_{unique_id}.xlsx'

        writer = pandas.ExcelWriter(file_name)

        dataframe.to_excel(writer, index=False)

        writer.close()
        cursor.close()
        dbsima.close()

        headerResponse = {
            'Content-Disposition': 'attachment; filename="' + file_name + '"'
        }

        background_task.add_task(os.remove, file_name)

        del dataframe

        return FileResponse(path=file_name, headers=headerResponse, filename=file_name)
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.post('/api/import/csv/sap3')
def import_csv_sap3(background_task: BackgroundTasks, file: UploadFile = File(...)):
    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        t0 = time.perf_counter()

        contents = file.file.read()

        with open(file.filename, 'wb') as f:
            f.write(contents)

        dataframe = pandas.read_csv(file.filename, sep=';', na_values='0', index_col=False, keep_default_na=False,
                                    dtype=str)
        dataframe.fillna(0, inplace=True)

        read_time = f"{time.perf_counter() - t0:.1f} seconds"

        cursor.execute("truncate table dev_real_sap3_tmp")
        cursor.commit()

        count = 0
        sql_statement = "BEGIN TRANSACTION \r\n"

        for index, df in dataframe.iterrows():
            count += 1

            dataframe.loc[index, 'amount_in_doc_curr'] = str(dataframe.loc[index, 'amount_in_doc_curr']).replace('.',
                                                                                                                 '').replace(
                ',', '.')
            dataframe.loc[index, 'dalam_jutaan'] = str(dataframe.loc[index, 'dalam_jutaan']).replace('.', '').replace(
                ',', '.')

            sql_statement += f"""
      insert into dev_real_sap3_tmp (kode_akun,kode_ba,kode_cc,posting_date,posting_period,no_dokumen,reference,assignment,amount_in_doc_curr,keterangan,trading_partner,
      dalam_jutaan,tenant,grouping,c3,c4,c5,c6,cc_baru,ba_baru,periode,cap_non_cap,refference_pm,reff_1,reff_2,portfolio,leveraging,digital,segmen,segmen_for_pl,dc,id_crm,id_ampm)
      values ('{df['kode_akun']}','{df['kode_ba']}','{df['kode_cc']}','{df['posting_date']}','{df['posting_period']}','{df['no_dokumen']}','{df['reference']}','{df['assignment']}',
      '{dataframe.loc[index, 'amount_in_doc_curr']}','{df['keterangan']}','{df['trading_partner']}','{dataframe.loc[index, 'dalam_jutaan']}','{df['tenant']}','{df['grouping']}','{df['c3']}',
      '{df['c4']}','{df['c5']}','{df['c6']}','{df['cc_baru']}','{df['ba_baru']}','{df['periode']}','{df['cap_non_cap']}','{df['refference_pm']}','{df['reff_1']}','{df['reff_2']}',
      '{df['portfolio']}','{df['leveraging']}','{df['digital']}','{df['segmen']}','{df['segmen_for_pl']}','{df['dc']}','{df['id_crm']}','{df['id_ampm']}')
      """

            if count % 100 == 0:
                sql_statement += "COMMIT TRANSACTION"
                cursor.execute(sql_statement)
                cursor.commit()
                sql_statement = "BEGIN TRANSACTION \r\n"

        if sql_statement != "BEGIN TRANSACTION \r\n":
            sql_statement += "COMMIT TRANSACTION"
            cursor.execute(sql_statement)
            cursor.commit()

        file.file.close()
        cursor.close()
        dbsima.close()

        background_task.add_task(os.remove, file.filename)

        return {"status": True, "message": "Import CSV berhasil", "read_time": read_time,
                "process_time": f"{time.perf_counter() - t0:.1f} seconds"}
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.post('/api/import/excel/sap3')
def import_excel_sap3(background_task: BackgroundTasks, file: UploadFile = File(...)):
    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        t0 = time.perf_counter()

        contents = file.file.read()

        with open(file.filename, 'wb') as f:
            f.write(contents)

        dataframe = pandas.read_excel(file.filename, na_values='0', index_col=False, keep_default_na=False, dtype=str)
        dataframe.fillna(0, inplace=True)

        read_time = f"{time.perf_counter() - t0:.1f} seconds"

        cursor.execute("truncate table dev_real_sap3_tmp")
        cursor.commit()

        count = 0
        sql_statement = "BEGIN TRANSACTION \r\n"

        for index, df in dataframe.iterrows():
            count += 1

            dataframe.loc[index, 'amount_in_doc_curr'] = str(dataframe.loc[index, 'amount_in_doc_curr']).replace('.',
                                                                                                                 '').replace(
                ',', '.')
            dataframe.loc[index, 'dalam_jutaan'] = str(dataframe.loc[index, 'dalam_jutaan']).replace('.', '').replace(
                ',', '.')

            sql_statement += f"""
      insert into dev_real_sap3_tmp (kode_akun,kode_ba,kode_cc,posting_date,posting_period,no_dokumen,reference,assignment,amount_in_doc_curr,keterangan,trading_partner,
      dalam_jutaan,tenant,grouping,c3,c4,c5,c6,cc_baru,ba_baru,periode,cap_non_cap,refference_pm,reff_1,reff_2,portfolio,leveraging,digital,segmen,segmen_for_pl,dc,id_crm,id_ampm)
      values ('{df['kode_akun']}','{df['kode_ba']}','{df['kode_cc']}','{df['posting_date']}','{df['posting_period']}','{df['no_dokumen']}','{df['reference']}','{df['assignment']}',
      '{dataframe.loc[index, 'amount_in_doc_curr']}','{df['keterangan']}','{df['trading_partner']}','{dataframe.loc[index, 'dalam_jutaan']}','{df['tenant']}','{df['grouping']}','{df['c3']}',
      '{df['c4']}','{df['c5']}','{df['c6']}','{df['cc_baru']}','{df['ba_baru']}','{df['periode']}','{df['cap_non_cap']}','{df['refference_pm']}','{df['reff_1']}','{df['reff_2']}',
      '{df['portfolio']}','{df['leveraging']}','{df['digital']}','{df['segmen']}','{df['segmen_for_pl']}','{df['dc']}','{df['id_crm']}','{df['id_ampm']}')
      """

            if count % 100 == 0:
                sql_statement += "COMMIT TRANSACTION"
                cursor.execute(sql_statement)
                cursor.commit()
                sql_statement = "BEGIN TRANSACTION \r\n"

        if sql_statement != "BEGIN TRANSACTION \r\n":
            sql_statement += "COMMIT TRANSACTION"
            cursor.execute(sql_statement)
            cursor.commit()

        file.file.close()
        cursor.close()
        dbsima.close()

        background_task.add_task(os.remove, file.filename)

        return {"status": True, "message": "Import Excel berhasil", "read_time": read_time,
                "process_time": f"{time.perf_counter() - t0:.1f} seconds"}
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.get('/api/template/perangkat-update')
def export_template_update_perangkat(background_task: BackgroundTasks, kode_regional: str | None = None,
                                     kode_area: str | None = None, kode_bm: str | None = None,
                                     kode_witel: str | None = None, kode_gedung: str | None = None):
    columns_perangkat = ["id", "kode_area", "kode_bm", "kode_witel", "kode_lokasi", "kode_gedung", "kode_lantai",
                         "kode_room", "kode_milik", "kode_jenis", "kode_kategori", "kode_subkategori", "kode_merk",
                         "jumlah", "satuan", "kapasitas", "satuan_kapasitas", "model", "no_seri", "tahun", "kondisi",
                         "waktu_pengadaan", "cek_birawa", "tanggal_cek", "keterangan",
                         "status_aktif", "updated_by", "updated_at"]

    where_perangkat = "where b.flag_aktif <> '3'"
    where_regional = "where a.kode_lokasi = '11'"
    where_area = "where a.kode_lokasi = '11'"
    where_bm = "where a.kode_lokasi = '11'"
    where_witel = "where a.kode_lokasi = '11'"
    where_location = "where a.flag_aktif = '1'"
    where_gedung = "where a.kode_lokasi = '11' and a.flag_aktif <> '3'"
    where_lantai = "where a.flag_aktif = '1'"
    where_ruang = "where a.flag_aktif = '1'"

    if kode_regional is not None and kode_regional != "" and kode_regional != "null":
        where_perangkat = where_perangkat + f"and b.kode_area = '{kode_regional}'"
        where_regional = where_regional + f"and a.kode_area = '{kode_regional}'"
        where_area = where_area + f"and a.kode_area = '{kode_regional}'"
        where_bm = where_bm + f"and a.kode_area = '{kode_regional}'"
        where_witel = where_witel + f"and b.kode_area = '{kode_regional}'"
        where_gedung = where_gedung + f"and a.kode_area = '{kode_regional}'"
        where_lantai = where_lantai + f"and b.kode_area = '{kode_regional}'"
        where_ruang = where_ruang + f"and b.kode_area = '{kode_regional}'"

    if kode_area is not None and kode_area != "" and kode_area != "null":
        where_perangkat = where_perangkat + f"and a.kode_fm = '{kode_area}'"
        where_bm = where_bm + f"and a.kode_fm = '{kode_area}'"
        where_witel = where_witel + f"and b.kode_fm = '{kode_area}'"
        where_gedung = where_gedung + f"and a.kode_fm = '{kode_area}'"
        where_lantai = where_lantai + f"and b.kode_fm = '{kode_area}'"
        where_ruang = where_ruang + f"and b.kode_fm = '{kode_area}'"

    if kode_bm is not None and kode_bm != "" and kode_bm != "null":
        where_perangkat = where_perangkat + f"and a.kode_bm = '{kode_bm}'"
        where_witel = where_witel + f"and b.kode_bm = '{kode_bm}'"
        where_location = where_location + f"and a.kode_bm = '{kode_bm}'"
        where_gedung = where_gedung + f"and a.kode_bm = '{kode_bm}'"
        where_lantai = where_lantai + f"and b.kode_bm = '{kode_bm}'"
        where_ruang = where_ruang + f"and b.kode_bm = '{kode_bm}'"

    if kode_witel is not None and kode_witel != "" and kode_witel != "null":
        where_perangkat = where_perangkat + f"and a.kode_witel = '{kode_witel}'"
        where_witel = where_witel + f"and b.kode_witel = '{kode_witel}'"
        where_gedung = where_gedung + f"and a.kode_witel = '{kode_witel}'"
        where_lantai = where_lantai + f"and b.kode_witel = '{kode_witel}'"
        where_ruang = where_ruang + f"and b.kode_witel = '{kode_witel}'"

    if kode_gedung is not None and kode_gedung != "" and kode_gedung != "null":
        where_perangkat = where_perangkat + f"and a.kode_gedung = '{kode_gedung}'"
        where_gedung = where_gedung + f"and a.kode_gedung = '{kode_gedung}'"
        where_lantai = where_lantai + f"and a.kode_gedung_sima = '{kode_gedung}'"
        where_ruang = where_ruang + f"and a.kode_gedung = '{kode_gedung}'"

    sql_statement_perangkat = f"""
  select a.id, a.kode_fm, a.kode_bm, a.kode_witel, a.kode_lokasi, a.kode_gedung, a.kode_lantai, a.kode_room, a.kode_milik, a.kode_jenis, a.kode_kategori, a.kode_subkategori, a.kode_merk, a.jumlah,
  a.satuan, a.kapasitas, a.satuan_kapasitas, a.model, a.no_seri, a.tahun, a.kondisi, a.tahun_pengadaan, a.is_ceklis, a.kondisi_terakhir, a.keterangan, isnull(a.status_aktif, 1) status_aktif, '' updated_by, '' updated_at
  from dev_am_perangkat a
  inner join am_gedung b on a.kode_gedung=b.kode_gedung and b.kode_lokasi='11'
  {where_perangkat}
  """

    columns_regional = ["kode_regional", "nama_regional"]

    sql_statement_regional = f"""
  select a.kode_area, a.nama
  from am_area a
  {where_regional}
  """

    columns_area = ["kode_area", "nama_area"]

    sql_statement_area = f"""
  select a.kode_fm, a.nama
  from am_fm a
  {where_area}
  """

    columns_bm = ["kode_bm", "nama"]

    sql_statement_bm = f"""
  select a.kode_bm, a.nama
  from am_bm a
  {where_bm}
  """

    columns_witel = ["kode_witel", "nama"]

    sql_statement_witel = f"""
  select a.kode_witel, a.nama
  from am_witel a
  where a.kode_witel = 'WT04'
  union all
  select a.kode_witel, a.nama
  from am_witel a
  inner join dev_am_perangkat_witel b on a.kode_witel=b.kode_witel
  {where_witel}
  group by a.kode_witel, a.nama
  """

    columns_location = ["kode_lokasi", "nama_lokasi", "penanggung_jawab"]

    sql_statement_location = f"""
  select a.id, a.nama_lokasi, a.penanggung_jawab
  from am_locations a
  {where_location}
  """

    columns_gedung = ["kode_gedung", "nama_gedung"]

    sql_statement_gedung = f"""
  select a.kode_gedung, a.nama_gedung
  from am_gedung a
  {where_gedung}
  """

    columns_lantai = ["kode_lantai", "nama_lantai", "kode_gedung"]

    sql_statement_lantai = f"""
  select a.id, a.nama_lantai, a.kode_gedung_sima kode_gedung
  from am_floors a
  inner join am_gedung b on a.kode_gedung_sima=b.kode_gedung and b.kode_lokasi='11'
  {where_lantai}
  order by a.kode_gedung_sima asc
  """

    columns_ruang = ["kode_ruang", "nama_ruang", "kode_lantai", "kode_gedung"]

    sql_statement_ruang = f"""
  select a.id, a.peruntukan, a.floor_id, a.kode_gedung
  from gsd_rooms a
  inner join am_gedung b on a.kode_gedung=b.kode_gedung and b.kode_lokasi='11'
  {where_ruang}
  order by a.kode_gedung asc
  """

    columns_jenis = ["kode_jenis", "nama_jenis"]

    sql_statement_jenis = f"""
  select a.jenis_id kode_jenis, a.nama_jenis
  from am_perangkat_jenis a
  """

    columns_kategori = ["kode_kategori", "nama_kategori", "kode_jenis"]

    sql_statement_kategori = f"""
  select a.kategori_id kode_kategori, a.nama_kategori, a.jenis_id kode_jenis
  from am_perangkat_kategori a
  """

    columns_subkategori = ["kode_kategori", "nama_kategori", "kode_jenis"]

    sql_statement_subkategori = f"""
  select a.sub_kategori_id kode_subkategori, a.nama_sub_kategori, a.kategori_id kode_kategori
  from am_perangkat_sub_kategori a
  """

    columns_merk = ["kode_merk", "nama_merk"]

    sql_statement_merk = f"""
  select a.kode_merk, a.nama_merk
  from am_perangkat_merk a
  """

    today = datetime.today()
    unique_id = today.strftime('%Y%m%d%H%M%S')
    file_name = f'UPDATE_PERANGKAT_{unique_id}.xlsx'

    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        writer = pandas.ExcelWriter(file_name, engine='openpyxl')

        cursor.execute(sql_statement_perangkat)

        dataframe_perangkat = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_perangkat)
        dataframe_perangkat.to_excel(writer, index=False, sheet_name='DATA_PERANGKAT')

        workbook = writer.book
        worksheet = workbook.active

        worksheet.merge_cells(start_row=1, start_column=30, end_row=6, end_column=35)
        worksheet['AD1'].alignment = Alignment(wrap_text=True)
        worksheet['AD1'].value = f"""
    Catatan :\n
    Mohon mengisi kolom updated_by dan updated_at\n
    jika melakukan perubahan data perangkat
    """

        worksheet['B1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['C1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['D1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['E1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['F1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['G1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['H1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['I1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['J1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['K1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['L1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')
        worksheet['M1'].fill = PatternFill(start_color='FFEA00', end_color='FFEA00', fill_type='solid')

        nsyyyymmdd = NamedStyle(name="nsyyyymmdd", number_format="YYYY-MM-DD")
        nsyyyymm = NamedStyle(name="nsyyyymm", number_format="YYYY-MM")

        for i in range(2, len(dataframe_perangkat) + 2):
            worksheet.cell(row=i, column=22).style = nsyyyymm
            worksheet.cell(row=i, column=24).style = nsyyyymmdd
            worksheet.cell(row=i, column=28).style = nsyyyymmdd

        cursor.execute(sql_statement_regional)

        dataframe_regional = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_regional)
        dataframe_regional.to_excel(writer, index=False, sheet_name='DATA_REGIONAL')

        cursor.execute(sql_statement_area)

        dataframe_area = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_area)
        dataframe_area.to_excel(writer, index=False, sheet_name='DATA_AREA')

        cursor.execute(sql_statement_bm)

        dataframe_bm = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_bm)
        dataframe_bm.to_excel(writer, index=False, sheet_name='DATA_BM')

        cursor.execute(sql_statement_witel)

        dataframe_witel = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_witel)
        dataframe_witel.to_excel(writer, index=False, sheet_name='DATA_WITEL')

        cursor.execute(sql_statement_location)

        dataframe_location = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_location)
        dataframe_location.to_excel(writer, index=False, sheet_name='DATA_LOCATION')

        cursor.execute(sql_statement_gedung)

        dataframe_gedung = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_gedung)
        dataframe_gedung.to_excel(writer, index=False, sheet_name='DATA_GEDUNG')

        cursor.execute(sql_statement_lantai)

        dataframe_lantai = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_lantai)
        dataframe_lantai.to_excel(writer, index=False, sheet_name='DATA_LANTAI')

        cursor.execute(sql_statement_ruang)

        dataframe_ruang = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_ruang)
        dataframe_ruang.to_excel(writer, index=False, sheet_name='DATA_RUANG')

        cursor.execute(sql_statement_jenis)

        dataframe_jenis = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_jenis)
        dataframe_jenis.to_excel(writer, index=False, sheet_name='DATA_JENIS')

        cursor.execute(sql_statement_kategori)

        dataframe_kategori = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_kategori)
        dataframe_kategori.to_excel(writer, index=False, sheet_name='DATA_KATEGORI')

        cursor.execute(sql_statement_subkategori)

        dataframe_subkategori = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_subkategori)
        dataframe_subkategori.to_excel(writer, index=False, sheet_name='DATA_SUBKATEGORI')

        cursor.execute(sql_statement_merk)

        dataframe_merk = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns_merk)
        dataframe_merk.to_excel(writer, index=False, sheet_name='DATA_MERK')

        headerResponse = {
            'Content-Disposition': 'attachment; filename="' + file_name + '"'
        }

        workbook.close()
        writer.close()
        cursor.close()
        dbsima.close()

        background_task.add_task(os.remove, file_name)

        del dataframe_perangkat
        del dataframe_regional
        del dataframe_area
        del dataframe_bm
        del dataframe_witel
        del dataframe_location
        del dataframe_gedung
        del dataframe_lantai
        del dataframe_ruang
        del dataframe_jenis
        del dataframe_kategori
        del dataframe_subkategori
        del dataframe_merk

        return FileResponse(path=file_name, headers=headerResponse, filename=file_name)
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.post('/api/import/perangkat-update')
def import_excel_perangkat_update(background_task: BackgroundTasks, file: UploadFile = File(...), nik: str = Form()):
    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        t0 = time.perf_counter()

        contents = file.file.read()

        with open(file.filename, 'wb') as f:
            f.write(contents)

        dataframe = pandas.read_excel(file.filename, na_values='0', index_col=False, keep_default_na=False, dtype=str,
                                      usecols='A:AB')
        dataframe.fillna(0, inplace=True)

        data_updated = dataframe.loc[(dataframe['updated_by'] != '') & (dataframe['updated_by'] != '')]

        read_time = f"{time.perf_counter() - t0:.1f} seconds"

        sql_statement = f"delete from dev_am_perangkat_upload where nik_upload = '{nik}'"
        cursor.execute(sql_statement)
        cursor.commit()

        count = 0
        sql_statement = "BEGIN TRANSACTION \r\n"

        for index, df in data_updated.iterrows():
            count += 1

            sql_statement += f"""
      insert into dev_am_perangkat_upload (id,kode_gedung,kode_room,kode_lantai,kode_merk,kode_lokasi,kode_fm,kode_bm,satuan,jumlah,kapasitas,satuan_kapasitas,model,no_seri,tahun,kondisi,
      kode_milik,keterangan,kode_jenis,kode_kategori,kode_subkategori,is_ceklis,kode_witel,kondisi_terakhir,tahun_pengadaan,status_aktif,updated_by,updated_at,nik_upload)
      values ('{df['id']}','{df['kode_gedung']}','{df['kode_room']}','{df['kode_lantai']}','{df['kode_merk']}','{df['kode_lokasi']}','{df['kode_area']}','{df['kode_bm']}','{df['satuan']}',
      '{df['jumlah']}','{df['kapasitas']}','{df['satuan_kapasitas']}','{df['model']}','{df['no_seri']}','{df['tahun']}','{df['kondisi']}','{df['kode_milik']}','{df['keterangan']}',
      '{df['kode_jenis']}','{df['kode_kategori']}','{df['kode_subkategori']}','{df['cek_birawa']}','{df['kode_witel']}','{df['tanggal_cek']}','{df['waktu_pengadaan']}',
      '{df['status_aktif']}','{df['updated_by']}','{df['updated_at']}','{nik}')
      """

            if count % 100 == 0:
                sql_statement += "COMMIT TRANSACTION"
                cursor.execute(sql_statement)
                cursor.commit()
                sql_statement = "BEGIN TRANSACTION \r\n"

        if sql_statement != "BEGIN TRANSACTION \r\n":
            sql_statement += "COMMIT TRANSACTION"
            cursor.execute(sql_statement)
            cursor.commit()

        file.file.close()
        cursor.close()
        dbsima.close()

        background_task.add_task(os.remove, file.filename)

        return {"status": True, "message": "Import Excel berhasil", "read_time": read_time,
                "process_time": f"{time.perf_counter() - t0:.1f} seconds"}
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.post('/api/scoring/linear-regression')
def calculate_linear_regression(var_model: Annotated[float, Body(embed=True)],
                                x: Annotated[List[float], Body(embed=True)], y: Annotated[List[int], Body(embed=True)]):
    slope, intercept, r, p, std_err = stats.linregress(x, y)

    def calculate(x):
        return slope * x + intercept

    def roundup(x):
        return math.ceil(x) + 0.000

    try:
        nilai = calculate(var_model)
        nilai_bs = roundup(calculate(var_model))

        if nilai_bs < 0:
            nilai_bs = nilai_bs * -1

        min_x = min(x)
        max_x = max(x)
        line_reg_start = calculate(min(x))
        line_reg_end = calculate(max(x))
        r_square = r ** 2

        data = {'nilai': nilai, 'nilai_bs': nilai_bs, 'var_model': var_model, 'r_squared': r_square, 'slope': slope,
                'intercept': intercept, 'r': r, 'p': p, 'standard_error': std_err, 'min_x': min_x, 'max_x': max_x,
                'line_reg_start': line_reg_start, 'line_ged_end': line_reg_end, 'x': x, 'y': y}

        return {"status": True, "message": "Success", "data": data}
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.get('/api/export/excel/notifikasi')
def export_excel_notifikasi(background_task: BackgroundTasks, filter_aset: Optional[str] = Query(default=None),
                            filter_dokumen: Optional[str] = Query(default=None),
                            filter_waktu: Optional[str] = Query(default=None),
                            nik: Optional[str] = Query(default=None)):
    filter_aset_list = filter_aset.split(",") if filter_aset else []
    filter_dokumen_list = filter_dokumen.split(",") if filter_dokumen else []
    filter_waktu_list = filter_waktu.split(",") if filter_waktu else []

    columns = ["ID ASET", "NAMA ASET", "JENIS DOKUMEN", "JENIS ASET", "TANGGAL AKHIR", "JATUH TEMPO", "REGIONAL"]
    data = []

    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        cursor.execute(f"""
    select a.kode_klp_menu tipe_user,b.kode_area
    from hakakses a
    inner join karyawan b on a.nik=b.nik and a.kode_lokasi=b.kode_lokasi
    where a.nik='{nik}' and a.kode_lokasi='11';               
    """)

        user = cursor.fetchone()

        where_regional = ""
        if user[0] == "AREA":
            where_regional = f"""and b.kode_area = '{user[1]}'"""

        jatuh_tempo = ""
        if "1m" in filter_waktu_list:
            if jatuh_tempo == "":
                jatuh_tempo = "a.tanggal_tenggat > DATEADD(MONTH, 1, GETDATE())"
            else:
                jatuh_tempo += "or a.tanggal_tenggat > DATEADD(MONTH, 1, GETDATE())"

        if "3m" in filter_waktu_list:
            if jatuh_tempo == "":
                jatuh_tempo = "a.tanggal_tenggat > DATEADD(MONTH, 3, GETDATE())"
            else:
                jatuh_tempo += "or a.tanggal_tenggat > DATEADD(MONTH, 3, GETDATE())"

        if "3w" in filter_waktu_list:
            if jatuh_tempo == "":
                jatuh_tempo = "a.tanggal_tenggat > DATEADD(WEEK, 3, GETDATE())"
            else:
                jatuh_tempo += "or a.tanggal_tenggat > DATEADD(WEEK, 3, GETDATE())"
        if "2w" in filter_waktu_list:
            if jatuh_tempo == "":
                jatuh_tempo = "a.tanggal_tenggat > DATEADD(WEEK, 2, GETDATE())"
            else:
                jatuh_tempo += "or a.tanggal_tenggat > DATEADD(WEEK, 2, GETDATE())"
        if "1w" in filter_waktu_list:
            if jatuh_tempo == "":
                jatuh_tempo = "a.tanggal_tenggat > DATEADD(WEEK, 1, GETDATE())"
            else:
                jatuh_tempo += "or a.tanggal_tenggat > DATEADD(WEEK, 1, GETDATE())"
        if "terlambat" in filter_waktu_list:
            if jatuh_tempo == "":
                jatuh_tempo = "a.tanggal_tenggat < GETDATE()"
            else:
                jatuh_tempo += "or a.tanggal_tenggat < GETDATE()"

        if "gedung" in filter_aset_list:
            list_dokumen = ""
            if "pbb" in filter_dokumen_list:
                if list_dokumen == "":
                    list_dokumen = "'PBB'"
                else:
                    list_dokumen += ",'PBB'"
            if "asuransi" in filter_dokumen_list:
                if list_dokumen == "":
                    list_dokumen = "'ASURANSI'"
                else:
                    list_dokumen += ",'ASURANSI'"
            if "tenant" in filter_dokumen_list:
                if list_dokumen == "":
                    list_dokumen = "'TENANT'"
                else:
                    list_dokumen += ",'TENANT'"

            cursor.execute(f"""
            select a.id, b.nama_gedung nama_aset, a.jenis_notif, a.jenis_aset, convert(varchar,a.tanggal_tenggat,103) deadline_date, CONVERT(VARCHAR,a.tanggal_tenggat,120) tanggal_tenggat,
            convert(varchar, GETDATE(), 23) tanggal_now, '' jatuh_tempo, c.nama
            from am_notifikasi a
            inner join am_gedung b on a.id=b.kode_gedung
            inner join am_area c on b.kode_area=c.kode_area and b.kode_lokasi=c.kode_lokasi
            where a.jenis_aset = 'GEDUNG' {where_regional} and a.jenis_notif in ({list_dokumen}) and ({jatuh_tempo})
            order by a.tanggal_tenggat asc
            """)

            for row in cursor.fetchall():
                date1 = datetime.strptime(row[5], '%Y-%m-%d')
                date2 = datetime.strptime(row[6], '%Y-%m-%d')

                if date1 < date2:
                    row[7] = "Sudah Habis"
                else:
                    interval = date1 - date2
                    years = interval.days
                    months = (interval.days % 365)
                    days = (interval.days % 365) % 30

                    keterangan = ""
                    if years > 0:
                        keterangan += f"{years} tahun "
                    if months > 0:
                        keterangan += f"{months} bulan "
                    if days > 0 or (years == 0 and months == 0):
                        keterangan += f"{days} hari"

                    row[7] = keterangan

                data.append([row[0], row[1], row[2], row[3], row[4], row[7], row[8]])

        if "kbm" in filter_aset_list:
            list_dokumen = ""
            if "pajak" in filter_dokumen_list:
                if list_dokumen == "":
                    list_dokumen = "'PAJAK'"
                else:
                    list_dokumen += ",'PAJAK'"
            if "asuransi" in filter_dokumen_list:
                if list_dokumen == "":
                    list_dokumen = "'ASURANSI'"
                else:
                    list_dokumen += ",'ASURANSI'"
            if "stnk" in filter_dokumen_list:
                if list_dokumen == "":
                    list_dokumen = "'STNK'"
                else:
                    list_dokumen += ",'STNK'"
            if "kontrak" in filter_dokumen_list:
                if list_dokumen == "":
                    list_dokumen = "'KONTRAK'"
                else:
                    list_dokumen += ",'KONTRAK'"

            cursor.execute(f"""
            select a.id, c.nopol nama_aset, a.jenis_notif, a.jenis_aset, convert(varchar,a.tanggal_tenggat,103) deadline_date, CONVERT(VARCHAR,a.tanggal_tenggat,120) tanggal_tenggat,
            convert(varchar, GETDATE(), 23) tanggal_now, '' jatuh_tempo, d.nama
            from am_notifikasi a
            inner join am_kbm c on a.id=c.id
            inner join am_gedung b on b.kode_gedung=c.id_gsd and b.kode_lokasi='11'
            inner join am_area d on b.kode_area=d.kode_area and b.kode_lokasi=d.kode_lokasi
            where a.jenis_aset = 'KBM' {where_regional} and a.jenis_notif in ({list_dokumen}) and ({jatuh_tempo})
            order by a.tanggal_tenggat asc
            """)

            for row in cursor.fetchall():
                date1 = datetime.strptime(row[5], '%Y-%m-%d')
                date2 = datetime.strptime(row[6], '%Y-%m-%d')

                if date1 < date2:
                    row[7] = "Sudah Habis"
                else:
                    interval = date1 - date2
                    years = interval.days
                    months = (interval.days % 365)
                    days = (interval.days % 365) % 30

                    keterangan = ""
                    if years > 0:
                        keterangan += f"{years} tahun "
                    if months > 0:
                        keterangan += f"{months} bulan "
                    if days > 0 or (years == 0 and months == 0):
                        keterangan += f"{days} hari"

                    row[7] = keterangan

                data.append([row[0], row[1], row[2], row[3], row[4], row[7], row[8]])

        if "perangkat" in filter_aset_list:
            list_dokumen = ""
            if "kondisi" in filter_dokumen_list:
                if list_dokumen == "":
                    list_dokumen = "'KONDISI'"
                else:
                    list_dokumen += ",'KONDISI'"

            cursor.execute(f"""
          select a.id, c.nama_perangkat nama_aset, a.jenis_notif, a.jenis_aset, convert(varchar,a.tanggal_tenggat,103) deadline_date, CONVERT(VARCHAR,a.tanggal_tenggat,120) tanggal_tenggat,
          convert(varchar, GETDATE(), 23) tanggal_now, '' jatuh_tempo, d.nama
          from am_notifikasi a
          inner join dev_am_perangkat c on a.id=c.id and isnull(c.status_aktif, '1')='1'
          inner join am_gedung b on b.kode_gedung=c.kode_gedung and b.kode_lokasi='11'
          inner join am_area d on b.kode_area=d.kode_area and b.kode_lokasi=d.kode_lokasi
          where a.jenis_aset = 'PERANGKAT' {where_regional} and a.jenis_notif in ({list_dokumen}) and ({jatuh_tempo})
          order by a.tanggal_tenggat asc
          """)

            for row in cursor.fetchall():
                date1 = datetime.strptime(row[5], '%Y-%m-%d')
                date2 = datetime.strptime(row[6], '%Y-%m-%d')

                if date1 < date2:
                    row[7] = "Sudah Habis"
                else:
                    interval = date1 - date2
                    years = interval.days
                    months = (interval.days % 365)
                    days = (interval.days % 365) % 30

                    keterangan = ""
                    if years > 0:
                        keterangan += f"{years} tahun "
                    if months > 0:
                        keterangan += f"{months} bulan "
                    if days > 0 or (years == 0 and months == 0):
                        keterangan += f"{days} hari"

                    row[7] = keterangan

                data.append([row[0], row[1], row[2], row[3], row[4], row[7], row[8]])

        dataframe = pandas.DataFrame.from_records(data, columns=columns)

        today = datetime.today()
        unique_id = today.strftime('%Y%m%d%H%M%S')

        file_name = f'DATA_NOTIFIKASI_{unique_id}.xlsx'

        writer = pandas.ExcelWriter(file_name)

        dataframe.to_excel(writer, index=False)

        writer.close()
        cursor.close()
        dbsima.close()

        headerResponse = {
            'Content-Disposition': 'attachment; filename="' + file_name + '"'
        }

        background_task.add_task(os.remove, file_name)

        del dataframe

        return FileResponse(path=file_name, headers=headerResponse, filename=file_name)
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.get('/api/export/excel/dashboard-kbm')
def export_excel_dashboardkbm(background_task: BackgroundTasks, filter_regional: Optional[str] = Query(default=None),
                              filter_area: Optional[str] = Query(default=None),
                              filter_pool: Optional[str] = Query(default=None),
                              filter_gedung: Optional[str] = Query(default=None),
                              filter_kepemilikan: Optional[str] = Query(default=None),
                              filter_tipe: Optional[str] = Query(default=None),
                              filter_search: Optional[str] = Query(default=None)):
    filter_regional_list = filter_regional.split(
        ",") if filter_regional is not None and filter_regional != "" and filter_regional != "null" else []
    filter_area_list = filter_area.split(
        ",") if filter_area is not None and filter_area != "" and filter_area != "null" else []
    filter_pool_list = filter_pool.split(
        ",") if filter_pool is not None and filter_pool != "" and filter_pool != "null" else []
    filter_gedung_list = filter_gedung.split(
        ",") if filter_gedung is not None and filter_gedung != "" and filter_gedung != "null" else []
    filter_kepemilikan_list = filter_kepemilikan.split(
        ",") if filter_kepemilikan is not None and filter_kepemilikan != "" and filter_kepemilikan != "null" else []
    filter_tipe_list = filter_tipe.split(
        ",") if filter_tipe is not None and filter_tipe != "" and filter_tipe != "null" else []

    def generate_where_in(filter):
        in_clause = ""

        if isinstance(filter, list):
            for key, value in enumerate(filter):
                if key == 0:
                    in_clause = f"'{value}'"
                else:
                    in_clause += f",'{value}'"
        return in_clause

    filter_regional = generate_where_in(filter_regional_list) if len(filter_regional_list) > 0 else ""
    in_regional = f"and a.reg_area in ({filter_regional})" if filter_regional != "" else ""
    filter_area = generate_where_in(filter_area_list) if len(filter_area_list) > 0 else ""
    in_area = f"and a.fm in ({filter_area})" if filter_area != "" else ""
    filter_pool = generate_where_in(filter_pool_list) if len(filter_pool_list) > 0 else ""
    in_pool = f"and a.pool_area in ({filter_pool})" if filter_pool != "" else ""
    filter_gedung = generate_where_in(filter_gedung_list) if len(filter_gedung_list) > 0 else ""
    in_gedung = f"and a.id_gsd in ({filter_gedung})" if filter_gedung != "" else ""
    filter_kepemilikan = generate_where_in(filter_kepemilikan_list) if len(filter_kepemilikan_list) > 0 else ""
    in_kepemilikan = f"and a.kepemilikan in ({filter_kepemilikan})" if filter_kepemilikan != "" else ""
    filter_tipe = generate_where_in(filter_tipe_list) if len(filter_tipe_list) > 0 else ""
    in_tipe = f"and a.tipe in ({filter_tipe})" if filter_tipe != "" else ""
    query_search = f"and (a.nopol like '%{filter_search}%' or a.nopol_awal like '%{filter_search}%' or a.lokasi_kbm like '%{filter_search}%')" if filter_search is not None and filter_search != "" and filter_search != "null" else ""

    columns = ["NOPOL", "NOPOL AWAL", "JENIS KENDARAAN", "MERK", "TIPE KENDARAAN", "TAHUN", "PERUNTUKAN", "LOKASI KBM",
               "REGIONAL", "KATEGORI MARKET", "USER", "NO ASET", "KONTRAK START", "KONTRAK END", "CUSTOMER", "AKTIVASI"]
    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        cursor.execute(f"""
    select a.nopol,a.nopol_awal,isnull(d.nama, '-') nama_jenis,isnull(e.nama, '-') nama_merk,isnull(c.nama, '-') nama_tipe,a.tahun,
    isnull(h.nama, '-') peruntukan,a.lokasi_kbm,isnull(b.nama,'-') nama_regional,isnull(j.nama,'-') nama_kategori,a.peruntukan kbm_user,
    isnull(a.no_aset,'-') no_aset,a.kontrak_start,a.kontrak_end,isnull(f.nama, '-') customer,a.aktivasi
    from am_kbm a
    left join am_area b on a.reg_area=b.kode_area
    left join am_kbm_type c on a.tipe=c.kode_type
    left join am_kbm_jenis d on a.jenis=d.kode_jenis
    left join am_kbm_merk e on a.merk=e.kode_merk
    left join am_kbm_cust f on a.costumer=f.kode_cust
    left join am_gedung g on a.id_gsd=g.kode_gedung and g.kode_lokasi='11'
    left join am_kbm_milik h on a.milik_kbm=h.kode_milik
    left join am_kbm_dispatcher i on a.dispatcher=i.kode_dispatcher
    left join am_kbm_group_cust j on a.group_cust=j.kode_group
    where a.aktivasi='ACTIVE' {in_regional} {in_area} {in_pool} {in_gedung} {in_kepemilikan} {in_tipe} {query_search}
    """)

        dataframe = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns)

        today = datetime.today()
        unique_id = today.strftime('%Y%m%d%H%M%S')

        file_name = f'DATA_KBM_{unique_id}.xlsx'

        writer = pandas.ExcelWriter(file_name)

        dataframe.to_excel(writer, index=False)

        writer.close()
        cursor.close()
        dbsima.close()

        headerResponse = {
            'Content-Disposition': 'attachment; filename="' + file_name + '"'
        }

        background_task.add_task(os.remove, file_name)

        del dataframe

        return FileResponse(path=file_name, headers=headerResponse, filename=file_name)
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.get('/api/export/excel/v2/dashboard-perangkat')
def export_excel_v2_perangkat(background_task: BackgroundTasks, filter_regional: Optional[str] = Query(default=None),
                              filter_area: Optional[str] = Query(default=None),
                              filter_bm: Optional[str] = Query(default=None),
                              filter_gedung: Optional[str] = Query(default=None),
                              filter_update: Optional[str] = Query(default=None),
                              filter_jenis: Optional[str] = Query(default=None),
                              filter_search: Optional[str] = Query(default=None)):
    filter_regional_list = filter_regional.split(
        ",") if filter_regional is not None and filter_regional != "" and filter_regional != "null" else []
    filter_area_list = filter_area.split(
        ",") if filter_area is not None and filter_area != "" and filter_area != "null" else []
    filter_bm_list = filter_bm.split(",") if filter_bm is not None and filter_bm != "" and filter_bm != "null" else []
    filter_gedung_list = filter_gedung.split(
        ",") if filter_gedung is not None and filter_gedung != "" and filter_gedung != "null" else []
    filter_jenis_list = filter_jenis.split(
        ",") if filter_jenis is not None and filter_jenis != "" and filter_jenis != "null" else []
    filter_update_list = filter_update.split(
        ",") if filter_update is not None and filter_update != "" and filter_update != "null" else []

    def generate_where_in(filter):
        in_clause = ""

        if isinstance(filter, list):
            for key, value in enumerate(filter):
                if key == 0:
                    in_clause = f"'{value}'"
                else:
                    in_clause += f",'{value}'"
        return in_clause

    filter_regional = generate_where_in(filter_regional_list) if len(filter_regional_list) > 0 else ""
    in_regional = f"and b.kode_area in ({filter_regional})" if filter_regional != "" else ""
    filter_area = generate_where_in(filter_area_list) if len(filter_area_list) > 0 else ""
    in_area = f"and b.kode_fm in ({filter_area})" if filter_area != "" else ""
    filter_bm = generate_where_in(filter_bm_list) if len(filter_bm_list) > 0 else ""
    in_bm = f"and b.kode_bm in ({filter_bm})" if filter_bm != "" else ""
    filter_gedung = generate_where_in(filter_gedung_list) if len(filter_gedung_list) > 0 else ""
    in_gedung = f"and a.kode_gedung in ({filter_gedung})" if filter_gedung != "" else ""
    filter_jenis = generate_where_in(filter_jenis_list) if len(filter_jenis_list) > 0 else ""
    in_jenis = f"and a.kode_jenis in ({filter_jenis})" if filter_jenis != "" else ""

    between_update = (
        f"FORMAT(a.tgl_input, 'yyyy-MM') between '{filter_update_list[0]}' and '{filter_update_list[1]}'"
        if filter_update_list is not None and len(filter_update_list) > 1
        else f"FORMAT(a.tgl_input, 'yyyy-MM') = '{filter_update_list[0]}'"
        if filter_update_list is not None and len(filter_update_list) == 1
        else ""
    )
    query_search = f"and (a.nama_perangkat like '%{filter_search}%' or a.id_perangkat like '%{filter_search}%' or l.nama_merk like '%{filter_search}%' or b.nama_gedung like '%{filter_search}%')" if filter_search is not None and filter_search != "" and filter_search != "null" else ""

    columns = ["no", "id_regional", "id_area", "nama_area", "id_bm", "nama_bm", "id_witel", "nama_witel", "id_location",
               "nama_lokasi", "id_gedung", "nama_gedung", "id_room", "id_lantai",
               "nama_lantai", "id_jenis", "nama_jenis", "id_kategori", "nama_kategori", "id_subkategori",
               "nama_subkategori", "nama_perangkat", "is_ceklis", "kode_merk", "nama_merk", "satuan", "jumlah",
               "kapasitas", "no_seri", "tipe", "tahun", "kondisi", "kode_milik", "nama_milik", "keterangan",
               "id_perangkat","id_perangkat_lama", "status_aktif", "tanggal_periksa", "nik_input", "updated_at"]

    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        cursor.execute(f"""
        select a.id no_perangkat, b.kode_area, b.kode_fm, j.nama nama_fm, a.kode_bm, k.nama nama_bm, b.kode_witel, i.nama nama_witel, a.kode_lokasi, c.nama_lokasi,
        a.kode_gedung, b.nama_gedung, a.kode_room, a.kode_lantai, e.nama_lantai, a.kode_jenis, f.nama_jenis,
        a.kode_kategori, g.nama_kategori, a.kode_subkategori, h.nama_sub_kategori, a.nama_perangkat, a.is_ceklis, a.kode_merk, l.nama_merk, a.satuan, a.jumlah,
        a.kapasitas, a.no_seri, a.model, a.tahun, a.kondisi, a.kode_milik, m.nama nama_milik, a.keterangan, concat(a.prefix,'-',a.indeks) id_perangkat,
        a.id_perangkat id_perangkat_lama, case when isnull(a.status_aktif, '1') = '1' then 'ACTIVE' else 'INACTIVE' end status_aktif, 
        a.kondisi_terakhir, a.nik_user, a.tgl_input
        from dev_am_perangkat a
        inner join am_gedung b on a.kode_gedung=b.kode_gedung and b.kode_lokasi='11'
        inner join am_locations c on a.kode_lokasi=c.id
        inner join gsd_rooms d on a.kode_room = d.id
        inner join am_floors e on a.kode_lantai=e.id
        left join am_perangkat_jenis f on a.kode_jenis=f.jenis_id
        left join am_perangkat_kategori g on a.kode_kategori=g.kategori_id
        left join am_perangkat_sub_kategori h on a.kode_subkategori=h.sub_kategori_id
        left join am_witel i on b.kode_witel=i.kode_witel
        left join am_fm j on a.kode_fm=j.kode_fm
        left join am_bm k on a.kode_bm=k.kode_bm
        left join am_perangkat_merk l on a.kode_merk=l.kode_merk
        left join am_milik m on a.kode_milik=m.kode_milik
        where e.flag_aktif <> '3' {in_regional} {in_area} {in_bm} {in_gedung} {in_jenis} {between_update} {query_search}
      """)

        dataframe = pandas.DataFrame.from_records(cursor.fetchall(), columns=columns)

        today = datetime.today()
        unique_id = today.strftime('%Y%m%d%H%M%S')

        file_name = f'DATA_PERANGKAT_{unique_id}.xlsx'

        writer = pandas.ExcelWriter(file_name)

        dataframe.to_excel(writer, index=False)

        writer.close()
        cursor.close()
        dbsima.close()

        headerResponse = {
            'Content-Disposition': 'attachment; filename="' + file_name + '"'
        }

        background_task.add_task(os.remove, file_name)

        del dataframe

        return FileResponse(path=file_name, headers=headerResponse, filename=file_name)
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.post('/api/performance/import/data-revenue')
def import_excel_performance_data_revenue(background_task: BackgroundTasks, file: UploadFile = File(...),
                                          nik: str = Form(), periode: str = Form()):
    t0 = time.perf_counter()

    today = datetime.today()
    tanggal_input = today.strftime('%Y-%m-%d %H:%M:%S')
    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        contents = file.file.read()

        with open(file.filename, 'wb') as f:
            f.write(contents)

        # skip baca baris ketiga
        # konversi kolom "ID GSD" menjadi string agar "0" didepan tidak dihilangkan
        dataframe = pandas.read_excel(file.filename, header=None, skiprows=[2], dtype={'ID GSD': str})

        read_time = f"{time.perf_counter() - t0:.1f} seconds"

        # Menghapus baris pertama dari dataframe
        df = dataframe.drop(0)
        # Melakukan iterasi untuk nilai i dalam rentang dari 1 hingga panjang baris pertama dataframe
        for i in range(1, len(df.iloc[0])):
            # Jika nilai pada baris pertama dan kolom i adalah NaN,
            # maka nilai tersebut akan diisi dengan nilai pada baris pertama dan kolom i-1 (sebelumnya)
            if pandas.isna(df.iloc[0, i]):
                df.iloc[0, i] = df.iloc[0, i - 1]

        data = []
        # Melakukan iterasi untuk nilai i dalam rentang dari 2 hingga panjang baris dari kolom pertama dataframe
        for i in range(2, len(df.iloc[0:, 0])):
            # Melakukan iterasi untuk nilai i dalam rentang dari 1 hingga panjang baris pertama dataframe
            for j in range(1, len(df.iloc[0])):
                data.append(dict(kode_gedung=df.iloc[i, 0], kode_revenue=df.iloc[0, j], kode_nilai=df.iloc[1, j],
                                 nilai=df.iloc[i, j]))

        sql_statement = f"delete from am_perf_gt_trans_revenue_tmp where nik_input='{nik}'"
        cursor.execute(sql_statement)
        cursor.commit()

        count = 0
        sql_statement = "BEGIN TRANSACTION \r\n"

        for i in range(0, len(data)):
            count += 1

            sql_statement += f"""
      insert into am_perf_gt_trans_revenue_tmp (kode_gedung,kode_revenue,kode_nilai,nilai,periode,nik_input,tanggal_input)
      values ('{data[i]['kode_gedung']}','{data[i]['kode_revenue']}','{data[i]['kode_nilai']}','{data[i]['nilai']}','{periode}','{nik}','{tanggal_input}')
      """

            if count % 100 == 0:
                sql_statement += "COMMIT TRANSACTION"
                cursor.execute(sql_statement)
                cursor.commit()
                sql_statement = "BEGIN TRANSACTION \r\n"

        if sql_statement != "BEGIN TRANSACTION \r\n":
            sql_statement += "COMMIT TRANSACTION"
            cursor.execute(sql_statement)
            cursor.commit()

        file.file.close()
        cursor.close()
        dbsima.close()

        background_task.add_task(os.remove, file.filename)

        return {"status": True, "message": "Import Excel berhasil", "read_time": read_time,
                "process_time": f"{time.perf_counter() - t0:.1f} seconds"}
    except Exception as ex:
        return {"status": False, "message": str(ex)}


@app.post('/api/performance/import/data-cost')
def import_excel_performance_data_cost(background_task: BackgroundTasks, file: UploadFile = File(...),
                                       nik: str = Form(), periode: str = Form()):
    t0 = time.perf_counter()

    today = datetime.today()
    tanggal_input = today.strftime('%Y-%m-%d %H:%M:%S')

    try:
        dbsima = connect_dbsima()
        cursor = dbsima.cursor()

        contents = file.file.read()

        with open(file.filename, 'wb') as f:
            f.write(contents)

        # skip baca baris ketiga
        # konversi kolom "ID GSD" menjadi string agar "0" didepan tidak dihilangkan
        dataframe = pandas.read_excel(file.filename, header=None, skiprows=[2], dtype={'ID GSD': str})

        read_time = f"{time.perf_counter() - t0:.1f} seconds"

        # Menghapus baris pertama dari dataframe
        df = dataframe.drop(0)
        # Melakukan iterasi untuk nilai i dalam rentang dari 1 hingga panjang baris pertama dataframe
        for i in range(1, len(df.iloc[0])):
            # Jika nilai pada baris pertama dan kolom i adalah NaN,
            # maka nilai tersebut akan diisi dengan nilai pada baris pertama dan kolom i-1 (sebelumnya)
            if pandas.isna(df.iloc[0, i]):
                df.iloc[0, i] = df.iloc[0, i - 1]

        data = []
        # Melakukan iterasi untuk nilai i dalam rentang dari 2 hingga panjang baris dari kolom pertama dataframe
        for i in range(2, len(df.iloc[0:, 0])):
            # Melakukan iterasi untuk nilai i dalam rentang dari 1 hingga panjang baris pertama dataframe
            for j in range(1, len(df.iloc[0])):
                data.append(dict(kode_gedung=df.iloc[i, 0], kode_cost=df.iloc[0, j], kode_cost_detail=df.iloc[1, j],
                                 nilai=df.iloc[i, j]))

        sql_statement = f"delete from am_perf_gt_trans_cost_tmp where nik_input='{nik}'"
        cursor.execute(sql_statement)
        cursor.commit()

        count = 0
        sql_statement = "BEGIN TRANSACTION \r\n"

        for i in range(0, len(data)):
            count += 1

            sql_statement += f"""
      insert into am_perf_gt_trans_cost_tmp (kode_gedung,kode_cost,kode_cost_detail,nilai,periode,nik_input,tanggal_input)
      values ('{data[i]['kode_gedung']}','{data[i]['kode_cost']}','{data[i]['kode_cost_detail']}','{data[i]['nilai']}','{periode}','{nik}','{tanggal_input}')
      """

            if count % 100 == 0:
                sql_statement += "COMMIT TRANSACTION"
                cursor.execute(sql_statement)
                cursor.commit()
                sql_statement = "BEGIN TRANSACTION \r\n"

        if sql_statement != "BEGIN TRANSACTION \r\n":
            sql_statement += "COMMIT TRANSACTION"
            cursor.execute(sql_statement)
            cursor.commit()

        file.file.close()

        background_task.add_task(os.remove, file.filename)

        return {"status": True, "message": "Import Excel berhasil", "read_time": read_time,
                "process_time": f"{time.perf_counter() - t0:.1f} seconds"}
    except Exception as ex:
        return {"status": False, "message": str(ex)}


class BaseRentBatch(BaseModel):
    kode_aset: str
    var_model: float
    y: List[float|int]
    x: List[float]


@app.post('/api/scoring/batch')
def calculate_linear_regression(assets: List[BaseRentBatch]):
    def calculate(sl, x, inter):
        return sl * x + inter

    def roundup(x):
        return math.ceil(x) + 0.000

    try:
        result = []
        for asset in assets:
            slope, intercept, r, p, std_err = stats.linregress(asset.x, asset.y)
            nilai = calculate(slope, asset.var_model, intercept)
            nilai_bs = roundup(calculate(slope, asset.var_model, intercept))

            if nilai_bs < 0:
                nilai_bs = nilai_bs * -1

            if nilai < 0:
                nilai = nilai * -1

            min_x = min(asset.x)
            max_x = max(asset.x)
            line_reg_start = calculate(slope, min(asset.x), intercept)
            line_reg_end = calculate(slope, max(asset.x), intercept)
            r_square = r ** 2

            data = dict([
                ("kode_aset", asset.kode_aset),
                ("nilai", nilai),
                ("nilai_bs", nilai_bs),
                ("var_model", asset.var_model),
                ("r_squared", r_square),
                ("slope", slope),
                ("intercept", intercept),
                ("r", r),
                ("p", p),
                ("standard_error", std_err),
                ("min_x", min_x),
                ("max_x", max_x),
                ("line_reg_start", line_reg_start),
                ("line_reg_end", line_reg_end),
                ("x", asset.x),
                ("y", asset.y)
            ])

            result.append(data)

        return {"status": True, "data": result}
    except Exception as ex:
        return {"status": False, "message": str(ex)}
